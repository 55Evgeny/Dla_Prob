import sys
import pdfplumber
import pandas as pd
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QVBoxLayout, QHBoxLayout, QWidget, QPushButton, QTableWidget, QTableWidgetItem, QFileDialog, QCheckBox, QHeaderView
)
from PyQt6.QtCore import Qt
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


class PDFViewerApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        self.setWindowTitle("PDF to Table Converter")
        self.setGeometry(100, 100, 800, 600)

        # Основной виджет и layout
        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)
        self.layout = QVBoxLayout(self.central_widget)

        # Кнопка для выбора PDF
        self.btn_open_pdf = QPushButton("Open PDF", self)
        self.btn_open_pdf.clicked.connect(self.open_pdf)
        self.layout.addWidget(self.btn_open_pdf)

        # Горизонтальный layout для чекбоксов
        self.checkbox_layout = QHBoxLayout()
        self.layout.addLayout(self.checkbox_layout)

        # Таблица для отображения данных
        self.table = QTableWidget(self)
        self.layout.addWidget(self.table)

        # Кнопка для экспорта выбранных столбцов
        self.btn_export = QPushButton("Export Selected Columns", self)
        self.btn_export.clicked.connect(self.export_selected_columns)
        self.layout.addWidget(self.btn_export)

        # Переменные для хранения данных
        self.data = None
        self.checkboxes = []

    def open_pdf(self):
        # Открываем диалог выбора файла
        file_path, _ = QFileDialog.getOpenFileName(self, "Open PDF File", "", "PDF Files (*.pdf)")
        if file_path:
            # Извлекаем таблицу из PDF
            self.data = self.extract_table_from_pdf(file_path)
            if self.data:
                # Отображаем данные в таблице
                self.display_data_in_table(self.data)

    def extract_table_from_pdf(self, file_path):
        # Используем pdfplumber для извлечения таблиц
        tables = []
        with pdfplumber.open(file_path) as pdf:
            for page in pdf.pages:
                # Извлекаем таблицу с текущей страницы
                table = page.extract_table()
                if table:
                    tables.extend(table)
        return tables

    def display_data_in_table(self, data):
        # Очищаем таблицу
        self.table.setRowCount(0)
        self.table.setColumnCount(0)

        # Устанавливаем количество строк и столбцов
        if data:
            self.table.setRowCount(len(data))
            self.table.setColumnCount(len(data[0]))

            # Заполняем таблицу данными
            for i, row in enumerate(data):
                for j, item in enumerate(row):
                    self.table.setItem(i, j, QTableWidgetItem(str(item) if item else ""))

            # Добавляем чекбоксы для выбора столбцов
            self.add_column_checkboxes(len(data[0]))

    def add_column_checkboxes(self, num_columns):
        # Очищаем предыдущие чекбоксы
        for checkbox in self.checkboxes:
            self.checkbox_layout.removeWidget(checkbox)
            checkbox.deleteLater()
        self.checkboxes.clear()

        # Создаем чекбоксы для каждого столбца
        for i in range(num_columns):
            checkbox = QCheckBox(f"Column {i + 1}", self)
            self.checkboxes.append(checkbox)
            self.checkbox_layout.addWidget(checkbox)

    def export_selected_columns(self):
        if self.data:
            # Получаем индексы выбранных столбцов
            selected_columns = [i for i, checkbox in enumerate(self.checkboxes) if checkbox.isChecked()]
            if selected_columns:
                # Создаем DataFrame с выбранными столбцами
                df = pd.DataFrame(self.data)
                df_selected = df.iloc[:, selected_columns]

                # Открываем диалог для выбора Excel-файла
                file_path, _ = QFileDialog.getSaveFileName(self, "Save File", "", "Excel Files (*.xlsx)")
                if file_path:
                    # Создаем Excel-файл
                    workbook = Workbook()
                    sheet = workbook.active

                    # Записываем заголовки столбцов
                    for col_num, header in enumerate(df_selected.columns, 1):
                        cell = sheet.cell(row=1, column=col_num, value=header)
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal='center')
                        cell.border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin')
                        )

                    # Записываем данные в Excel
                    for row in dataframe_to_rows(df_selected, index=False, header=False):
                        sheet.append(row)

                    # Применяем форматирование к данным
                    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                        for cell in row:
                            cell.alignment = Alignment(horizontal='center')
                            cell.border = Border(
                                left=Side(style='thin'),
                                right=Side(style='thin'),
                                top=Side(style='thin'),
                                bottom=Side(style='thin')
                            )

                    # Сохраняем файл
                    workbook.save(file_path)


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = PDFViewerApp()
    window.show()
    sys.exit(app.exec())
