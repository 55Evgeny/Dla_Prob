
import sys
import pdfplumber
import pandas as pd
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QVBoxLayout, QHBoxLayout, QWidget, QPushButton, QTableWidget, QTableWidgetItem, QFileDialog, QCheckBox
)
from PyQt6.QtCore import Qt
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


class PDFViewerApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        self.setWindowTitle("PDF to Table Converter")
        self.setGeometry(100, 100, 800, 600)

        # Основной виджет и layout
        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)
        self.layout = QVBoxLayout(self.central_widget)

        # Кнопка для выбора PDF
        self.btn_open_pdf = QPushButton("Открыть PDF", self)
        self.btn_open_pdf.clicked.connect(self.open_pdf)
        self.layout.addWidget(self.btn_open_pdf)

        # Горизонтальный layout для чекбоксов
        self.checkbox_layout = QHBoxLayout()
        self.layout.addLayout(self.checkbox_layout)

        # Таблица для отображения данных
        self.table = QTableWidget(self)
        self.layout.addWidget(self.table)

        # Кнопка для экспорта выбранных столбцов
        self.btn_export = QPushButton("Перенести выбранные столбцы в excel", self)
        self.btn_export.clicked.connect(self.export_selected_columns)
        self.layout.addWidget(self.btn_export)

        # Переменные для хранения данных
        self.data = None
        self.checkboxes = []

    def open_pdf(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "Open PDF File", "", "PDF Files (*.pdf)")
        if file_path:
            self.data = self.extract_table_from_pdf(file_path)
            if self.data is not None:
                self.display_data_in_table(self.data)

    def extract_table_from_pdf(self, file_path):
        tables = []
        try:
            with pdfplumber.open(file_path) as pdf:
                for page in pdf.pages:
                    table = page.extract_table()
                    if table:
                        tables.extend(table)
            if tables:
                headers = tables[0]
                data = tables[1:]
                df = pd.DataFrame(data, columns=headers)
                return df
            else:
                print("No tables found in the PDF.")
                return None
        except Exception as e:
            print(f"Error extracting tables: {e}")
            return None

    def display_data_in_table(self, data):
        self.table.setRowCount(0)
        self.table.setColumnCount(0)

        if data is not None and not data.empty:
            self.table.setRowCount(data.shape[0])
            self.table.setColumnCount(data.shape[1])

            self.table.setHorizontalHeaderLabels(data.columns)

            for i, row in data.iterrows():
                for j, item in enumerate(row):
                    self.table.setItem(i, j, QTableWidgetItem(str(item) if pd.notna(item) else ""))

            self.add_column_checkboxes(data.shape[1])
        else:
            print("No data to display.")

    def add_column_checkboxes(self, num_columns):
        for checkbox in self.checkboxes:
            self.checkbox_layout.removeWidget(checkbox)
            checkbox.deleteLater()
        self.checkboxes.clear()

        for i in range(num_columns):
            checkbox = QCheckBox(f"Column {i + 1}", self)
            self.checkboxes.append(checkbox)
            self.checkbox_layout.addWidget(checkbox)

    def export_selected_columns(self):
        if self.data is not None and not self.data.empty:
            selected_columns = [i for i, checkbox in enumerate(self.checkboxes) if checkbox.isChecked()]
            if selected_columns:
                df_selected = self.data.iloc[:, selected_columns]

                file_path, _ = QFileDialog.getSaveFileName(self, "Save File", "", "Excel Files (*.xlsx)")
                if file_path:
                    workbook = Workbook()
                    sheet = workbook.active

                    for col_num, header in enumerate(df_selected.columns, 1):
                        cell = sheet.cell(row=1, column=col_num, value=header)
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal='center')
                        cell.border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin')
                        )

                    for row in dataframe_to_rows(df_selected, index=False, header=False):
                        sheet.append(row)

                    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                        for cell in row:
                            cell.alignment = Alignment(horizontal='center')
                            cell.border = Border(
                                left=Side(style='thin'),
                                right=Side(style='thin'),
                                top=Side(style='thin'),
                                bottom=Side(style='thin')
                            )

                    workbook.save(file_path)
            else:
                print("No columns selected.")
        else:
            print("No data to export.")


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = PDFViewerApp()
    window.show()
    sys.exit(app.exec())
