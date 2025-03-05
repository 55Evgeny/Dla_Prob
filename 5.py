import sys
import pandas as pd
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QVBoxLayout, QHBoxLayout, QWidget, QPushButton, QTableWidget, QTableWidgetItem, QFileDialog, QCheckBox
)
from PyQt6.QtCore import Qt
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import pdfplumber


class PDFViewerApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        self.setWindowTitle("PDF to Table Converter")
        self.setGeometry(100, 100, 800, 600)

        # Основной виджет и layout
        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)
        self.layout = QVBoxLayout(self.central_widget)

        # Кнопка для выбора PDF
        self.btn_open_pdf = QPushButton("Open PDF", self)
        self.btn_open_pdf.clicked.connect(self.open_pdf)
        self.layout.addWidget(self.btn_open_pdf)

        # Горизонтальный layout для чекбоксов
        self.checkbox_layout = QHBoxLayout()
        self.layout.addLayout(self.checkbox_layout)

        # Таблица для отображения данных
        self.table = QTableWidget(self)
        self.layout.addWidget(self.table)

        # Кнопка для экспорта выбранных столбцов
        self.btn_export = QPushButton("Export Selected Columns", self)
        self.btn_export.clicked.connect(self.export_selected_columns)
        self.layout.addWidget(self.btn_export)

        # Переменные для хранения данных
        self.data = None
        self.checkboxes = []

    def open_pdf(self):
        # Открываем диалог выбора файла
        file_path, _ = QFileDialog.getOpenFileName(self, "Open PDF File", "", "PDF Files (*.pdf)")
        if file_path:
            # Извлекаем таблицу из PDF с помощью pdfplumber
            self.data = self.extract_table_from_pdf(file_path)
            if self.data is not None:
                # Отображаем данные в таблице
                self.display_data_in_table(self.data)

    def extract_table_from_pdf(self, file_path):
        try:
            # Используем pdfplumber для извлечения таблиц
            with pdfplumber.open(file_path) as pdf:
                tables = []
                for page in pdf.pages:
                    # Извлекаем таблицы с текущей страницы
                    table = page.extract_table()
                    if table:
                        tables.extend(table)

                if tables:
                    # Преобразуем список таблиц в DataFrame
                    df = pd.DataFrame(tables[1:], columns=tables[0])

                    # Ручная обработка данных
                    df = self.process_data(df)
                    return df
                else:
                    print("No tables found in the PDF.")
                    return None
        except Exception as e:
            print(f"Error extracting tables: {e}")
            return None

    def process_data(self, df):
        # Объединяем данные в нужные столбцы
        # Например, если столбец 5 содержит описание работ, объединяем все строки в одну ячейку
        if df.shape[1] > 5:  # Убедимся, что столбец 5 существует
            df[4] = df.iloc[:, 4:].apply(lambda row: " ".join(row.dropna().astype(str)), axis=1)

        # Удаляем лишние столбцы
        df = df.iloc[:, :10]  # Оставляем только первые 10 столбцов

        # Возвращаем обработанный DataFrame
        return df

    def display_data_in_table(self, data):
        # Очищаем таблицу
        self.table.setRowCount(0)
        self.table.setColumnCount(0)

        # Устанавливаем количество строк и столбцов
        if not data.empty:
            self.table.setRowCount(data.shape[0])
            self.table.setColumnCount(data.shape[1])

            # Заполняем таблицу данными
            for i, row in data.iterrows():
                for j, item in enumerate(row):
                    self.table.setItem(i, j, QTableWidgetItem(str(item) if pd.notna(item) else ""))

            # Добавляем чекбоксы для выбора столбцов
            self.add_column_checkboxes(data.shape[1])

    def add_column_checkboxes(self, num_columns):
        # Очищаем предыдущие чекбоксы
        for checkbox in self.checkboxes:
            self.checkbox_layout.removeWidget(checkbox)
            checkbox.deleteLater()
        self.checkboxes.clear()

        # Создаем чекбоксы для каждого столбца
        for i in range(num_columns):
            checkbox = QCheckBox(f"Column {i + 1}", self)
            self.checkboxes.append(checkbox)
            self.checkbox_layout.addWidget(checkbox)

    def export_selected_columns(self):
        if self.data is not None:
            # Получаем индексы выбранных столбцов
            selected_columns = [i for i, checkbox in enumerate(self.checkboxes) if checkbox.isChecked()]
            if selected_columns:
                # Создаем DataFrame с выбранными столбцами
                df_selected = self.data.iloc[:, selected_columns]

                # Открываем диалог для выбора Excel-файла
                file_path, _ = QFileDialog.getSaveFileName(self, "Save File", "", "Excel Files (*.xlsx)")
                if file_path:
                    # Создаем Excel-файл
                    workbook = Workbook()
                    sheet = workbook.active

                    # Записываем заголовки столбцов
                    for col_num, header in enumerate(df_selected.columns, 1):
                        cell = sheet.cell(row=1, column=col_num, value=header)
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal='center')
                        cell.border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin')
                        )

                    # Записываем данные в Excel
                    for row in dataframe_to_rows(df_selected, index=False, header=False):
                        sheet.append(row)

                    # Применяем форматирование к данным
                    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                        for cell in row:
                            cell.alignment = Alignment(horizontal='center')
                            cell.border = Border(
                                left=Side(style='thin'),
                                right=Side(style='thin'),
                                top=Side(style='thin'),
                                bottom=Side(style='thin')
                            )

                    # Сохраняем файл
                    workbook.save(file_path)


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = PDFViewerApp()
    window.show()
    sys.exit(app.exec())
